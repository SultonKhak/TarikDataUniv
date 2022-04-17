import requests
from requests.adapters import HTTPAdapter
from requests.exceptions import ConnectionError
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import mysql.connector

mydb = mysql.connector.connect(
  host="localhost",
  user="root",
  password="",
  database="kampus")

headers = {
    'user-agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36'
}


github_adapter = HTTPAdapter(max_retries=3)

session = requests.Session()
session.mount('https://api-frontend.kemdikbud.go.id/v2/detail_pt_prodi', github_adapter)

book = openpyxl.load_workbook('Data API Universitas.xlsx')

sheet = book['Baru']
m_row = sheet.max_row

workbook = Workbook()
sheett = workbook.active

def tarik_data(row):
    for i in range(row, m_row + 1):
        A = sheet.cell(row=i, column=1)
        B = sheet.cell(row=i, column=2)
        C= sheet.cell(row=i, column=3)
        D = sheet.cell(row=i, column=4)
        F = sheet.cell(row=i, column=5)
        J = sheet.cell(row=i, column=10)
        univ = A.value
        api = B.value
        kode_pt = C.value
        if sheet.cell(row=i, column=4) is None:
            status_pt = '-'
        else:
            status_pt = D.value

        if sheet.cell(row=i, column=5) is None:
            akreditasi_pt = '-'
        else:
            akreditasi_pt = F.value
        wilayah = J.value

        url = 'https://api-frontend.kemdikbud.go.id/v2/detail_pt_prodi/' + str(api)
        data = requests.get(url).json()
        print(url)

        try:
            respone = session.get(url, headers=headers)
        except ConnectionError as ce:
            print(ce)

        while respone.status_code < 200:
            respone = requests.get(url)
            print(respone.status_code)

        if (respone.text == ''):
            data = []
        else:
            data = respone.json()

        if data:
            panjang = len(data)

            for i in range(panjang):
                kode_prodi = data[i]["kode_prodi"]
                nama_lembaga = data[i]["nm_lemb"]
                status_prodi = data[i]["stat_prodi"]
                jenjang = data[i]["jenjang"]
                if(data[i]["akreditas"] is None):
                    akreditasi = '-'
                else:
                    akreditasi = data[i]["akreditas"]

                if(len(data[i]["rasio_list"]) == 5):
                    if (data[i]["rasio_list"][0]["semester"] == '20211'):
                        jumlah_mhs_20191 = data[i]["rasio_list"][4]["mahasiswa"]
                        jumlah_mhs_20192 = data[i]["rasio_list"][3]["mahasiswa"]
                        jumlah_mhs_20201 = data[i]["rasio_list"][2]["mahasiswa"]
                        jumlah_mhs_20202 = data[i]["rasio_list"][1]["mahasiswa"]
                        jumlah_mhs_20211 = data[i]["rasio_list"][0]["mahasiswa"]
                        # sheett['A1'] = "kode_prodi"
                        # sheett['B1'] = "nama_lembaga"
                        # sheett['C1'] = "status_prodi"
                        # sheett['D1'] = "jenjang"
                        # sheett['E1'] = "akreditasi"
                        # sheett['F1'] = "jumlah_mhs_20191"
                        # sheett['G1'] = "jumlah_mhs_20192"
                        # sheett['H1'] = "jumlah_mhs_20201"
                        # sheett['I1'] = "jumlah_mhs_20202"
                        # sheett['J1'] = "jumlah_mhs_20211"
                        # sheett['A' + str(i + 2)].value = kode_prodi
                        # sheett['B' + str(i + 2)].value = nama_lembaga
                        # sheett['C' + str(i + 2)].value = status_prodi
                        # sheett['D' + str(i + 2)].value = jenjang
                        # sheett['E' + str(i + 2)].value = akreditasi
                        # sheett['F' + str(i + 2)].value = jumlah_mhs_20191
                        # sheett['G' + str(i + 2)].value = jumlah_mhs_20192
                        # sheett['H' + str(i + 2)].value = jumlah_mhs_20201
                        # sheett['I' + str(i + 2)].value = jumlah_mhs_20202
                        # sheett['J' + str(i + 2)].value = jumlah_mhs_20211
                        print("sucses")
                    else:
                        print("code tidak sesuai dengan case PDDIKTI")

                mycursor = mydb.cursor()
                sql = "INSERT INTO pddikti.new_table(nama_pt,kode_pt,status_pt,akreditasi_pt,wilayah_dan_kabupaten,nama_prodi,kode_prodi,status_prodi,jenjang,akreditasi,jumlah_mhs_20191,jumlah_mhs_20192,jumlah_mhs_20201,jumlah_mhs_20202,jumlah_mhs_20211) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                print(sql)
                val = (univ, kode_pt, status_pt, akreditasi_pt, wilayah, nama_lembaga, kode_prodi, status_prodi, jenjang, akreditasi, jumlah_mhs_20191, jumlah_mhs_20192, jumlah_mhs_20201, jumlah_mhs_20202, jumlah_mhs_20211)
                print(val)
                mycursor.execute(sql, val)

                mydb.commit()

        # workbook.save(filename=univ+".xlsx")

        for i in range(len(data)):
            sheett['A' + str(i + 2)].value = ""
            sheett['B' + str(i + 2)].value = ""
            sheett['C' + str(i + 2)].value = ""
            sheett['D' + str(i + 2)].value = ""
            sheett['E' + str(i + 2)].value = ""
            sheett['F' + str(i + 2)].value = ""
            sheett['G' + str(i + 2)].value = ""
            sheett['H' + str(i + 2)].value = ""
            sheett['I' + str(i + 2)].value = ""
            sheett['J' + str(i + 2)].value = ""

tarik_data(2)
